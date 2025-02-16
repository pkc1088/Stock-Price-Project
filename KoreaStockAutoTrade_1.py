import os
import requests
import json
import datetime
import time
import yaml
import openpyxl
from openpyxl import load_workbook

with open('config.yaml', encoding='UTF-8') as f:
    _cfg = yaml.load(f, Loader=yaml.FullLoader)
APP_KEY = _cfg['APP_KEY']
APP_SECRET = _cfg['APP_SECRET']
ACCESS_TOKEN = ""
CANO = _cfg['CANO']
ACNT_PRDT_CD = _cfg['ACNT_PRDT_CD']
DISCORD_WEBHOOK_URL = _cfg['DISCORD_WEBHOOK_URL']
URL_BASE = _cfg['URL_BASE']


def send_message(msg):
    """디스코드 메세지 전송"""
    now = datetime.datetime.now()
    message = {"content": f"[{now.strftime('%Y-%m-%d %H:%M:%S')}] {str(msg)}"}
    requests.post(DISCORD_WEBHOOK_URL, data=message)
    print(message)


def get_access_token():
    """토큰 발급"""
    headers = {"content-type": "application/json"}
    body = {"grant_type": "client_credentials",
            "appkey": APP_KEY,
            "appsecret": APP_SECRET}
    PATH = "oauth2/tokenP"
    URL = f"{URL_BASE}/{PATH}"
    res = requests.post(URL, headers=headers, data=json.dumps(body))
    ACCESS_TOKEN = res.json()["access_token"]
    return ACCESS_TOKEN


def hashkey(datas):
    """암호화"""
    PATH = "uapi/hashkey"
    URL = f"{URL_BASE}/{PATH}"
    headers = {
        'content-Type': 'application/json',
        'appKey': APP_KEY,
        'appSecret': APP_SECRET,
    }
    res = requests.post(URL, headers=headers, data=json.dumps(datas))
    hashkey = res.json()["HASH"]
    return hashkey


import pandas as pd

def sort_excel_sheet_by_columns(file_name, sheet_name):
    """
    주어진 시트에서 SoldOut이 비어있는 행은 StockCode와 BuyPrice 순으로 정렬하고,
    SoldOut에 문자가 있는 행은 맨 끝으로 이동하는 함수.

    Args:
        file_name (str): 엑셀 파일 경로
        sheet_name (str): 정렬할 시트 이름
    """
    # 엑셀 파일 불러오기
    df = pd.read_excel(file_name, sheet_name=sheet_name)

    # SoldOut 열이 비어있는 데이터 필터링
    df_empty_soldout = df[df['SoldOut'].isnull() | (df['SoldOut'] == '')]

    # SoldOut 열에 문자가 있는 데이터 필터링
    df_non_empty_soldout = df[df['SoldOut'].notnull() & (df['SoldOut'] != '')]

    # StockCode와 BuyPrice 순으로 오름차순 정렬
    df_sorted = df_empty_soldout.sort_values(by=['StockCode', 'BuyPrice'], ascending=[True, True])

    # 두 데이터프레임 결합
    df_final = pd.concat([df_sorted, df_non_empty_soldout], ignore_index=True)

    # 엑셀 파일에 다시 저장
    with pd.ExcelWriter(file_name, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
        df_final.to_excel(writer, sheet_name=sheet_name, index=False)


def record_buy(code, qty, price):
    """매수 기록을 엑셀 파일에 추가하고 정렬"""
    buy_price = int(price)  # 예시로 응답에서 매수 가격 가져오기
    timestamp = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    # 엑셀 파일을 로드
    excel_file = 'auto_trade.xlsx'
    wb = load_workbook(excel_file)

    if 'BuyRecords' not in wb.sheetnames:
        ws = wb.create_sheet('BuyRecords')
    else:
        ws = wb['BuyRecords']

    # 매수 기록 추가
    ws.append([timestamp, code, buy_price, qty])

    # 엑셀 파일에 저장
    wb.save(excel_file)

    # BuyRecords 시트를 StockCode와 BuyPrice에 따라 정렬
    sort_excel_sheet_by_columns(excel_file, 'BuyRecords')

def record_sell(code, qty, price):
    """매도 기록을 엑셀 파일에 추가하고 정렬"""
    sell_price = int(price)  # 예시로 응답에서 매도 가격 가져오기
    timestamp = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    # 엑셀 파일을 로드
    excel_file = 'auto_trade.xlsx'
    wb = load_workbook(excel_file)

    if 'SellRecords' not in wb.sheetnames:
        ws = wb.create_sheet('SellRecords')
    else:
        ws = wb['SellRecords']

    # 매도 기록 추가
    ws.append([timestamp, code, sell_price, qty])

    # BuyRecords 시트를 업데이트
    if 'BuyRecords' in wb.sheetnames:
        buy_ws = wb['BuyRecords']

        # StockCode가 가장 먼저 일치하는 열에 SoldOut 체크
        for row in range(2, buy_ws.max_row + 1):
            stock_code_cell = buy_ws[f'B{row}']  # StockCode 열
            quantity_cell = buy_ws[f'D{row}']  # Quantity 열
            if stock_code_cell.value == code:
                buy_qty = quantity_cell.value
                if buy_qty > 0:  # 매수 수량이 남아있을 경우
                    quantity_cell.value = buy_qty - qty if buy_qty >= qty else 0  # 매도 수량 차감
                    if quantity_cell.value == 0:
                        buy_ws[f'E{row}'] = 'O'  # SoldOut 열에 'O' 입력
                        break  # 가장 먼저 일치한 열을 찾았으므로 반복 종료

    # 엑셀 파일에 저장
    wb.save(excel_file)

    # BuyRecords 시트를 StockCode와 BuyPrice에 따라 정렬
    sort_excel_sheet_by_columns('auto_trade.xlsx', 'BuyRecords')

def get_current_price(code):
    """현재가 조회 및 엑셀에 저장"""
    PATH = "uapi/domestic-stock/v1/quotations/inquire-price"
    URL = f"{URL_BASE}/{PATH}"
    headers = {"Content-Type": "application/json",
               "authorization": f"Bearer {ACCESS_TOKEN}",
               "appKey": APP_KEY,
               "appSecret": APP_SECRET,
               "tr_id": "FHKST01010100"}
    params = {
        "fid_cond_mrkt_div_code": "J",
        "fid_input_iscd": code,
    }
    res = requests.get(URL, headers=headers, params=params)
    current_price = int(res.json()['output']['stck_prpr'])

    # 엑셀 파일을 읽어와서 업데이트
    excel_file = 'auto_trade.xlsx'
    sheet_name = 'TargetPrices'
    # 엑셀 파일을 로드
    wb = load_workbook(excel_file)
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(sheet_name)
    else:
        ws = wb[sheet_name]

    # 주식 코드에 맞는 행을 찾음 (리스트로 변환하여 비교)
    stock_codes = [cell.value for cell in ws['A'] if cell.value is not None]
    if int(code) in stock_codes:
        # 해당 코드의 인덱스를 찾음
        idx = stock_codes.index(int(code)) + 1
        # 해당 행의 'CurrentPrice' 열을 현재 가격으로 업데이트
        ws[f'B{idx}'] = current_price
    else:
        # 주식 코드가 없을 경우 새로운 행을 추가
        new_row = [int(code), current_price, 0]
        ws.append(new_row)

    # 'TargetBuyPrice' 열에 수식 추가
    for row in range(2, ws.max_row + 1):
        ws[f'C{row}'] = f'==IFERROR(IF(VLOOKUP(A{row}, BuyRecords!B:E, 4, FALSE) = "O", B{row}, VLOOKUP(A{row}, BuyRecords!B:C, 2, FALSE) * 0.99), B{row})'

    # 'TargetSellPrice' 열에 수식 추가
    for row in range(2, ws.max_row + 1):
        ws[f'D{row}'] = f'=IFERROR(VLOOKUP(A{row},BuyRecords!B:C,2,FALSE)*1.01,"")'

    # 엑셀 파일에 저장
    wb.save(excel_file)
    return current_price

import win32com.client
def get_target_price(code):
    try:
        # Excel 애플리케이션 열기
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False  # Excel 창을 보이지 않게 설정

        # 엑셀 파일 열기
        wb = excel.Workbooks.Open(os.path.abspath('auto_trade.xlsx'))

        # TargetPrices 시트 선택
        ws = wb.Sheets('TargetPrices')

        # 시트의 모든 데이터를 순회하면서 StockCode를 찾기
        for row in range(2, ws.UsedRange.Rows.Count + 1):  # 첫 번째 행은 헤더로 건너뜀
            stock_code = ws.Cells(row, 1).Value  # StockCode
            current_price = ws.Cells(row, 2).Value # CurrentPrice
            target_buy_price = ws.Cells(row, 3).Value  # TargetBuyPrice
            target_sell_price = ws.Cells(row, 4).Value # TargetSellPrice
            # 값을 제대로 반환하는지 확인
            if int(stock_code) == int(code):
                print(f"({datetime.datetime.now()})StockCode: {stock_code}, CurrentPrice: {current_price}, TargetBuyPrice: {target_buy_price}, target_sell_price: {target_sell_price}")
                wb.Close(SaveChanges=False)  # 엑셀 파일 닫기
                excel.Quit()  # Excel 종료
                return target_buy_price, target_sell_price  # 계산된 TargetBuyPrice 값을 반환
        wb.Close(SaveChanges=False)  # 엑셀 파일 닫기
        excel.Quit()  # Excel 종료
        return None  # 코드가 없으면 None 리턴

    except Exception as e:
        print(f"Error occurred: {e}")
        return None

def create_auto_trade_excel():
    """auto_trade.xlsx 파일 생성"""
    if not os.path.exists('auto_trade.xlsx'):
        # 초기 데이터프레임 생성
        columns_target = ['StockCode', 'CurrentPrice', 'TargetBuyPrice', 'TargetSellPrice']
        df_target = pd.DataFrame(columns=columns_target)

        # 매수 기록용 데이터프레임 생성
        columns_buy = ['Timestamp', 'StockCode', 'BuyPrice', 'Quantity', 'SoldOut']
        df_buy = pd.DataFrame(columns=columns_buy)

        # 매도 기록용 데이터프레임 생성
        columns_sell = ['Timestamp', 'StockCode', 'SellPrice', 'Quantity']
        df_sell = pd.DataFrame(columns=columns_sell)

        # Excel 파일에 저장
        with pd.ExcelWriter('auto_trade.xlsx') as writer:
            df_target.to_excel(writer, sheet_name='TargetPrices', index=False)
            df_buy.to_excel(writer, sheet_name='BuyRecords', index=False)
            df_sell.to_excel(writer, sheet_name='SellRecords', index=False)

        send_message("auto_trade.xlsx 파일이 생성되었습니다.")
    else:
        send_message("auto_trade.xlsx 파일이 이미 존재합니다.")

def get_stock_balance():
    """주식 잔고조회"""
    PATH = "uapi/domestic-stock/v1/trading/inquire-balance"
    URL = f"{URL_BASE}/{PATH}"
    headers = {"Content-Type": "application/json",
               "authorization": f"Bearer {ACCESS_TOKEN}",
               "appKey": APP_KEY,
               "appSecret": APP_SECRET,
               "tr_id": "TTTC8434R",
               "custtype": "P",
               }
    params = {
        "CANO": CANO,
        "ACNT_PRDT_CD": ACNT_PRDT_CD,
        "AFHR_FLPR_YN": "N",
        "OFL_YN": "",
        "INQR_DVSN": "02",
        "UNPR_DVSN": "01",
        "FUND_STTL_ICLD_YN": "N",
        "FNCG_AMT_AUTO_RDPT_YN": "N",
        "PRCS_DVSN": "01",
        "CTX_AREA_FK100": "",
        "CTX_AREA_NK100": ""
    }
    res = requests.get(URL, headers=headers, params=params)
    stock_list = res.json()['output1']
    evaluation = res.json()['output2']
    stock_dict = {}
    send_message(f"====주식 보유잔고====")
    for stock in stock_list:
        if int(stock['hldg_qty']) > 0:
            stock_dict[stock['pdno']] = stock['hldg_qty']
            send_message(f"{stock['prdt_name']}({stock['pdno']}): {stock['hldg_qty']}주")
            time.sleep(0.1)
    send_message(f"주식 평가 금액: {evaluation[0]['scts_evlu_amt']}원")
    time.sleep(0.1)
    send_message(f"평가 손익 합계: {evaluation[0]['evlu_pfls_smtl_amt']}원")
    time.sleep(0.1)
    send_message(f"총 평가 금액: {evaluation[0]['tot_evlu_amt']}원")
    time.sleep(0.1)
    send_message(f"=================")
    return stock_dict


def get_balance():
    """현금 잔고조회"""
    PATH = "uapi/domestic-stock/v1/trading/inquire-psbl-order"
    URL = f"{URL_BASE}/{PATH}"
    headers = {"Content-Type": "application/json",
               "authorization": f"Bearer {ACCESS_TOKEN}",
               "appKey": APP_KEY,
               "appSecret": APP_SECRET,
               "tr_id": "TTTC8908R",
               "custtype": "P",
               }
    params = {
        "CANO": CANO,
        "ACNT_PRDT_CD": ACNT_PRDT_CD,
        "PDNO": "005930",
        "ORD_UNPR": "65500",
        "ORD_DVSN": "01",
        "CMA_EVLU_AMT_ICLD_YN": "Y",
        "OVRS_ICLD_YN": "Y"
    }
    res = requests.get(URL, headers=headers, params=params)
    cash = res.json()['output']['ord_psbl_cash']
    send_message(f"주문 가능 현금 잔고: {cash}원")
    return int(cash)


def buy(code, buy_price, qty):
    """주식 시장가 매수"""
    PATH = "uapi/domestic-stock/v1/trading/order-cash"
    URL = f"{URL_BASE}/{PATH}"
    data = {
        "CANO": CANO,
        "ACNT_PRDT_CD": ACNT_PRDT_CD,
        "PDNO": code,
        "ORD_DVSN": "01",
        "ORD_QTY": str(int(qty)),
        "ORD_UNPR": "0",
    }
    headers = {"Content-Type": "application/json",
               "authorization": f"Bearer {ACCESS_TOKEN}",
               "appKey": APP_KEY,
               "appSecret": APP_SECRET,
               "tr_id": "TTTC0802U",
               "custtype": "P",
               "hashkey": hashkey(data)
               }
    res = requests.post(URL, headers=headers, data=json.dumps(data))
    if res.json()['rt_cd'] == '0':
        send_message(f"[매수 성공]{str(res.json())}")
        record_buy(code, int(qty), buy_price)
        return True
    else:
        send_message(f"[매수 실패]{str(res.json())}")
        return False


def sell(code, sell_price, qty):
    """주식 시장가 매도"""
    PATH = "uapi/domestic-stock/v1/trading/order-cash"
    URL = f"{URL_BASE}/{PATH}"
    data = {
        "CANO": CANO,
        "ACNT_PRDT_CD": ACNT_PRDT_CD,
        "PDNO": code,
        "ORD_DVSN": "01",
        "ORD_QTY": str(int(qty)),
        "ORD_UNPR": "0",
    }
    headers = {"Content-Type": "application/json",
               "authorization": f"Bearer {ACCESS_TOKEN}",
               "appKey": APP_KEY,
               "appSecret": APP_SECRET,
               "tr_id": "TTTC0801U",
               "custtype": "P",
               "hashkey": hashkey(data)
               }
    res = requests.post(URL, headers=headers, data=json.dumps(data))
    if res.json()['rt_cd'] == '0':
        send_message(f"[매도 성공]{str(res.json())}")
        record_sell(code, int(qty), sell_price)
        return True
    else:
        send_message(f"[매도 실패]{str(res.json())}")
        return False

def get_order_price(code):
    """체결된 주문의 가격을 조회"""
    PATH = "uapi/domestic-stock/v1/trading/inquire-daily-ccld"
    URL = f"{URL_BASE}/{PATH}"
    headers = {
        "Content-Type": "application/json",
        "authorization": f"Bearer {ACCESS_TOKEN}",
        "appKey": APP_KEY,
        "appSecret": APP_SECRET,
        "tr_id": "TTTC8001R",  # 주문 상세 조회 API 트랜잭션 ID
        "tr_cont": "",
        "custtype" : "P",
    }
    params = {
        "CANO": CANO,
        "ACNT_PRDT_CD": ACNT_PRDT_CD,
        "INQR_STRT_DT": "20240831",#datetime.datetime.now().strftime('%Y%m%d'),  # 조회 시작 날짜
        "INQR_END_DT": datetime.datetime.now().strftime('%Y%m%d'),    # 조회 종료 날짜 (오늘)
        "SLL_BUY_DVSN_CD": "00",  # 전체 조회
        "INQR_DVSN": "00",        # 역순
        "PDNO": code,             # 종목 코드
        "CCLD_DVSN": "01",        # 체결만
        "ORD_GNO_BRNO": "",
        "ODNO" : "",
        "INQR_DVSN_3" : "00",
        "INQR_DVSN_1" : "",
        "CTX_AREA_FK100": "",     # Context ID
        "CTX_AREA_NK100": "",     # Context ID
    }
    res = requests.get(URL, headers=headers, params=params)
    if res.json()['rt_cd'] == '0':
        # 주문 체결 내역에서 체결된 가격을 가져옴
        order_details = res.json().get('output1', [])
        print(res.json())
        if order_details:
            filled_price = int(order_details[0]['stck_prpr'])  # 체결된 가격
            return filled_price
    else:
        send_message(f"[체결 가격 조회 실패] {str(res.json())}")
    return None

# 자동매매 시작
try:
    ACCESS_TOKEN = get_access_token()
    create_auto_trade_excel()

    bought_prices = {}
    symbol_list = ["005930"]  # 매수 희망 종목 리스트
    total_cash = get_balance()  # 보유 현금 조회
    stock_dict = get_stock_balance()  # 보유 주식 조회

    send_message("===국내 주식 자동매매 프로그램을 시작합니다===")
    while True:
        t_now = datetime.datetime.now()
        t_9 = t_now.replace(hour=9, minute=0, second=0, microsecond=0)
        t_start = t_now.replace(hour=9, minute=5, second=0, microsecond=0)
        t_sell = t_now.replace(hour=15, minute=15, second=0, microsecond=0)
        t_exit = t_now.replace(hour=15, minute=20, second=0, microsecond=0)
        today = datetime.datetime.today().weekday()
        if today == 5 or today == 6:  # 토요일이나 일요일이면 자동 종료
            send_message("주말이므로 프로그램을 종료합니다.")
            break

        if t_start < t_now < t_sell:  # AM 09:05 ~ PM 03:15 : 매수, 매도
            for sym in symbol_list:
                target_buy_price, target_sell_price = get_target_price(sym)
                current_price = get_current_price(sym)
                if target_buy_price is not None and current_price is not None:
                    # 매수 목표가 이하일 때 매수
                    if int(target_buy_price) >= int(current_price):
                        buy_qty = 1  # 매수할 수량 초기화
                        if (total_cash // current_price) > 0:
                            send_message(f"{sym} 매수 목표가 달성({target_buy_price} > {current_price}) 매수를 시도합니다.")
                            result = buy(sym, target_buy_price, buy_qty)
                            if result:
                                stock_dict = get_stock_balance()  # 주식 잔고 업데이트
                                total_cash = get_balance()  # 현금 잔고 업데이트
                                target_buy_price, target_sell_price = get_target_price(sym)

                # 매도 목표가 이상일 때 매도
                if sym in stock_dict and target_sell_price is not None and current_price is not None:
                    if int(target_sell_price) <= int(current_price):
                        sell_qty = 1  # 매도할 수량 초기화
                        if int(stock_dict[sym]) > 0:
                            send_message(f"{sym} 매도 목표가 달성({target_sell_price} <= {current_price}) 매도를 시도합니다.")
                            result = sell(sym, target_sell_price, sell_qty)
                            if result:
                                stock_dict = get_stock_balance()  # 주식 잔고 업데이트
                                total_cash = get_balance()  # 현금 잔고 업데이트
                                target_buy_price, target_sell_price = get_target_price(sym)

                time.sleep(1)
            time.sleep(1)
            if t_now.minute == 30 and t_now.second <= 5:
                get_stock_balance()
                time.sleep(5)
        if t_exit < t_now:  # PM 03:20 ~ :프로그램 종료
            send_message("프로그램을 종료합니다.")
            break

except Exception as e:
    send_message(f"[오류 발생] {e}")
    import traceback
    send_message(traceback.format_exc())
    time.sleep(1)